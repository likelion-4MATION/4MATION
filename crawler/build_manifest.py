"""crawl_manifest.csv 병합 생성.

입력:
  kdic_필수페이지_URL매핑.csv        — 필수 30건
  분석필요태깅_포함여부분석.xlsx      — '포함' 판정분만 병합

출력 컬럼: url, business_function, sub_category, page_type,
          coverage(전체/안내부), variant, source(필수/분석포함)

컬럼명이 정확히 일치하지 않아도 자동 탐지 후 매핑 결과를 출력하니
반드시 미리보기를 눈으로 확인할 것.

사용: python build_manifest.py 필수.csv 분석.xlsx -o crawl_manifest.csv
"""

from __future__ import annotations

import argparse
import csv
import re
import sys

OUT_COLS = ["url", "business_function", "sub_category", "page_type",
            "coverage", "variant", "source"]

# 자동 탐지 키워드 → 출력 컬럼
COL_HINTS = {
    "url": ("url", "주소", "링크"),
    "business_function": ("business", "업무", "기능"),
    "sub_category": ("sub", "세부", "카테고리", "분류"),
    "page_type": ("page_type", "유형", "타입", "page_name"),
    "coverage": ("coverage", "수집범위", "범위", "안내부"),
    "variant": ("variant", "분기", "판"),
}
INCLUDE_HINTS = ("판정", "포함여부", "포함", "결과")


def detect(headers: list[str]) -> dict[str, str]:
    mapping = {}
    for out_col, hints in COL_HINTS.items():
        for h in headers:
            if any(k in str(h).lower() for k in hints):
                mapping[out_col] = h
                break
    return mapping


def norm_row(raw: dict, mapping: dict[str, str], source: str) -> dict | None:
    url = str(raw.get(mapping.get("url", ""), "") or "").strip()
    if not url.startswith("http"):
        return None
    row = {c: str(raw.get(mapping.get(c, ""), "") or "").strip() for c in OUT_COLS[:-1]}
    row["url"] = url
    row["source"] = source
    return row


def read_csv(path: str) -> list[dict]:
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        mapping = detect(headers)
        print(f"[csv] 컬럼 매핑: {mapping}")
        return [r for raw in reader if (r := norm_row(raw, mapping, "필수"))]


def read_xlsx(path: str) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []
    for ws in wb:
        data = list(ws.iter_rows(values_only=True))
        if not data:
            continue
        headers = [str(h or "") for h in data[0]]
        mapping = detect(headers)
        inc_col = next((h for h in headers
                        if any(k in h for k in INCLUDE_HINTS)), None)
        # 판정컬럼('예비판정' 등)이 '판' 힌트로 variant에 오매핑되는 것 방지
        mapping = {k: v for k, v in mapping.items() if v != inc_col}
        print(f"[xlsx:{ws.title}] 컬럼 매핑: {mapping} · 판정컬럼: {inc_col}")
        for values in data[1:]:
            raw = dict(zip(headers, values))
            verdict = str(raw.get(inc_col, "") or "") if inc_col else ""
            # '포함' 판정만 병합, '제외/보류/대기' 등은 스킵
            if inc_col and not re.search(r"포함", verdict):
                continue
            if inc_col and re.search(r"제외|보류|대기|미포함", verdict):
                continue
            if r := norm_row(raw, mapping, "분석포함"):
                # xlsx 표기 정규화 — 업무명 번호 프리픽스 제거('5. 채무조정 안내'),
                # '포함(안내부)' 판정은 coverage로 전파 (내용 창작 아님, 판정 텍스트 유래)
                r["business_function"] = re.sub(r"^\s*\d+\.\s*", "", r["business_function"])
                if "안내부" in verdict and not r["coverage"]:
                    r["coverage"] = "안내부"
                rows.append(r)
    return rows


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("required_csv")
    ap.add_argument("analysis_xlsx", nargs="?", default=None,
                    help="'포함' 판정 병합용 xlsx (없으면 필수 CSV만으로 확정)")
    ap.add_argument("-o", "--out", default="crawl_manifest.csv")
    args = ap.parse_args()

    rows = read_csv(args.required_csv)
    if args.analysis_xlsx:
        rows += read_xlsx(args.analysis_xlsx)
    else:
        print("[xlsx] 분석필요 xlsx 미지정 — 필수 CSV만으로 확정 (추후 병합 필요)")
    seen, uniq = set(), []
    for r in rows:
        if r["url"] not in seen:
            seen.add(r["url"])
            uniq.append(r)
    if not uniq:
        sys.exit("병합 결과 0건 — 컬럼 매핑 출력을 확인할 것")

    with open(args.out, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=OUT_COLS)
        w.writeheader()
        w.writerows(uniq)

    print(f"\n총 {len(uniq)}건 → {args.out}")
    print("미리보기:")
    for r in uniq[:5]:
        print(f"  [{r['source']}] {r['business_function']} | {r['url']}")


if __name__ == "__main__":
    main()
